import tkinter as tk
from tkinter import filedialog
import openpyxl
import random

preguntas = []
nombre_archivo = ""

# Ventana principal
ventana = tk.Tk()
ventana.title('Mezclador de preguntas')

# Función para cargar el archivo Excel seleccionado
def cargar_excel():
    global nombre_archivo
    nombre_archivo = filedialog.askopenfilename(filetypes=[("Archivos Excel", "*.xlsx;*.xls")])
    
    if nombre_archivo:
        # Nombre del archivo en el label
        label_archivo.config(text=f"Archivo seleccionado: {nombre_archivo}")
        
        # abrimos el excel
        wb = openpyxl.load_workbook(nombre_archivo)
        ws = wb.active

        for row in range(2, ws.max_row + 1):
            pregunta = ws.cell(row=row, column=1).value
            opciones = [ws.cell(row=row, column=col).value for col in range(2, ws.max_column + 1)]
            # print(opciones)
            random.shuffle(opciones)
            # print(opciones)
            global preguntas
            preguntas.append((pregunta, opciones))

def definir_rangos():
    global rangos
    rangos = []
    cantidad_preguntas = [int(entry.get()) for entry in entries]
    inicio = 1
    for cantidad in cantidad_preguntas:
        fin = inicio + cantidad - 1
        rangos.append((inicio, fin))
        inicio = fin + 1
    print("rangos definidos")

def mezclar_preguntas():
    for inicio, fin in rangos:
        sublista = preguntas[inicio-1:fin]
        random.shuffle(sublista)
        preguntas[inicio-1:fin] = sublista
    print("preguntas mezcladas")

def guardar_preguntas():
    wb_out = openpyxl.Workbook()
    ws_out = wb_out.active

    for idx, (pregunta, opciones) in enumerate(preguntas, start=1):
        ws_out.cell(row=idx, column=1, value=pregunta)
        for i, opcion in enumerate(opciones, start=2):
            ws_out.cell(row=idx, column=i, value=opcion)

    nombre_guardar = nombre_archivo[:-5] + "_ALEATORIZADO.xlsx"
    wb_out.save(nombre_guardar)
    print("doc guardado")

# Boton cargar archivo excel
boton_cargar = tk.Button(ventana, text="Seleccionar Excel", command=cargar_excel)
boton_cargar.grid(row=1, column=0, padx=10, pady=5, columnspan=2)

# Label nombre archivo
label_archivo = tk.Label(ventana, text="Archivo seleccionado: ")
label_archivo.grid(row=2, column=0, padx=10, pady=5, columnspan=2)

# # Crear el widget Text para mostrar el contenido del archivo Excel
# texto_excel = tk.Text(ventana, wrap=tk.NONE)
# texto_excel.grid(row=3, column=0, padx=10, pady=5, columnspan=2)

etiqueta_rangos = tk.Label(ventana, text='COMPLETE LA CANTIDAD DE PREGUNTAS POR AREA')
etiqueta_rangos.grid(row=4, column=0, padx=10, pady=5, columnspan=2)

areas = [
    "aritmetica",
    "algebra",
    "geometria",
    "trigonometria",
    "fisica",
    "quimica",
    "biologia y anatomia",
    "psicologia y filosofia",
    "geografia",
    "historia",
    "educacion civica",
    "economia",
    "comunicacion",
    "literatura",
    "razonamiento matematico",
    "razonamiento verbal",
    "ingles",
    "quechua y aimara",
]

entries = []

# Etiqueta y campo de entrada para cada rango
for i, area in enumerate(areas, start=1):
    etiqueta_rango = tk.Label(ventana, text=f'Área {area}:')
    etiqueta_rango.grid(row=i+4, column=0, padx=10, pady=5)

    entry = tk.Entry(ventana, width=5)
    entry.grid(row=i+4, column=1, padx=10, pady=5)
    entries.append(entry)

# Botón para definir los rangos
boton_definir_rangos = tk.Button(ventana, text='Definir rangos', command=definir_rangos)
boton_definir_rangos.grid(row=len(areas) + 5, column=0, columnspan=2, padx=10, pady=10)

# Botón para mezclar las preguntas
boton_mezclar_preguntas = tk.Button(ventana, text='Mezclar preguntas', command=mezclar_preguntas)
boton_mezclar_preguntas.grid(row=len(areas) + 6, column=0, columnspan=2, padx=10, pady=10)

# Botón para guardar las preguntas mezcladas en un archivo Excel
boton_guardar_preguntas = tk.Button(ventana, text='Guardar preguntas mezcladas', command=guardar_preguntas)
boton_guardar_preguntas.grid(row=len(areas) + 7, column=0, columnspan=2, padx=10, pady=10)

ventana.mainloop()
